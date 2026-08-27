"""Fill the OF-288 Excel template (ref/OF288 Excel Blank Template.xlsx)
instead of the fillable PDF. Reuses paycheck8_parser + of288_filler's
build_groups/allocate_rows unchanged — those are already format-agnostic,
they just produce structured data (which incident got which hours on which
date). Only the "how do we write it down" step differs here.

The Excel template has 8 grid rows per column (vs. 7 on the PDF) and no
fixed-height AcroForm box to fight for font size — so this sidesteps the
clipping/rect-enlargement problem the PDF needed.

Important: the "Hours" and "Total Hours" cells (including item 17's grand
total) are LIVE FORMULAS in the template, not blank cells — Hours is
computed from Start/Stop via a military-time-difference formula, and each
column's total is a SUM() over its own Hours column, with item 17 summing
all four column totals. This code only ever writes Mo/Day/Start/Stop/the
H-T-E flag; it never touches an Hours or Total Hours cell, so those
formulas stay intact and Excel/LibreOffice recomputes them correctly on
open. (Verified: our own computed hours always equals Stop-minus-Start for
every row we produce, by construction of allocate_rows — the template's
formula and our own number necessarily agree.)

Cell positions were derived programmatically from the template's merged-
cell structure and cross-checked against two real filled examples (a Rev
10/2015 OF-288: "Moyer OF28 PP15.xlsx" and "Moyers OF288 PP16.xlsx"), not
guessed — except one transcription slip (column C's total-hours cell)
caught by a self-verification pass before this ever touched real data.
"""
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

from of288_filler import allocate_rows

FONT_NAME = "Calibri"
FONT_SIZE = 12
FONT_BOLD = False
GRID_ROWS = 8
GRID_FIRST_ROW = 20
YEAR_TOTAL_ROW = 28
COLUMN_LETTERS = ["A", "B", "C", "D"]
BLOCK_BASE_COL = {"A": "A", "B": "G", "C": "N", "D": "W"}

TYPE_OF_EMPLOYMENT_CELLS = {"Casual": "J4", "Federal": "M4", "Other": "O4"}

# "Same as Column" mark cells for a continuation column — e.g. Column B's
# "Same as Column [ ]A" checkbox is J8 (right before the static "A" label
# at K8). Mirrors the PDF version's same_as_checkboxes.
SAME_AS_CHECKBOXES = {
    "B": {"A": "J8"},
    "C": {"A": "Q8", "B": "S8"},
    "D": {"A": "Y8", "B": "AA8", "C": "AC8"},
}

TOP_CELLS = {
    "1_hired_at": "R2",
    "2_employee_common_identifier": "A4",
    "4_hiring_unit_name": "R4",
    "5_name": "A6",
    "6_hiring_unit_phone": "Q6",
    "7_hiring_unit_fax": "X6",
}

REMARKS_CELL = "A38"
ITEM17_TOTAL_CELL = "AB29"


def _offset(base_letter, n):
    return get_column_letter(column_index_from_string(base_letter) + n)


def _build_column_map(base_letter):
    """Per-column-block cell map, derived by a fixed offset from that
    block's base column (A/G/N/W). Offsets aren't uniform across blocks
    because Day/Start/Stop are visually merged to different widths per
    block — these exact values were verified against the template's own
    merge structure and two real filled examples, not assumed to follow a
    formula.
    """
    if base_letter == "A":
        subfield_offsets = {"9": 0, "10": 0, "11": 2, "12": 0, "13": 2, "14": 3, "15": 0}
        grid_offsets = {"mo": 0, "day": 1, "start": 2, "stop": 3, "hours": 4, "flag": 5}
    elif base_letter == "G":
        subfield_offsets = {"9": 0, "10": 0, "11": 2, "12": 0, "13": 2, "14": 4, "15": 0}
        grid_offsets = {"mo": 0, "day": 1, "start": 2, "stop": 4, "hours": 5, "flag": 6}
    elif base_letter == "N":
        subfield_offsets = {"9": 0, "10": 0, "11": 3, "12": 0, "13": 3, "14": 5, "15": 0}
        grid_offsets = {"mo": 0, "day": 1, "start": 3, "stop": 5, "hours": 7, "flag": 8}
    elif base_letter == "W":
        subfield_offsets = {"9": 0, "10": 0, "11": 2, "12": 0, "13": 2, "14": 4, "15": 0}
        grid_offsets = {"mo": 0, "day": 1, "start": 2, "stop": 4, "hours": 6, "flag": 7}
    else:
        raise ValueError(base_letter)

    rows = {"8": 10, "9": 12, "10": 14, "11": 14, "12": 16, "13": 16, "14": 16, "15": 18}
    meta_cells = {
        "incident_name": f"{base_letter}{rows['8']}",
        "incident_order_number": f"{_offset(base_letter, subfield_offsets['9'])}{rows['9']}",
        "fire_code": f"{_offset(base_letter, subfield_offsets['10'])}{rows['10']}",
        "resource_request_number": f"{_offset(base_letter, subfield_offsets['11'])}{rows['11']}",
        "position_code": f"{_offset(base_letter, subfield_offsets['12'])}{rows['12']}",
        "ad_class": f"{_offset(base_letter, subfield_offsets['13'])}{rows['13']}",
        "ad_rate": f"{_offset(base_letter, subfield_offsets['14'])}{rows['14']}",
        "accounting_code": f"{_offset(base_letter, subfield_offsets['15'])}{rows['15']}",
    }
    grid = []
    for r in range(GRID_ROWS):
        row = GRID_FIRST_ROW + r
        grid.append({
            "mo": f"{_offset(base_letter, grid_offsets['mo'])}{row}",
            "day": f"{_offset(base_letter, grid_offsets['day'])}{row}",
            "start": f"{_offset(base_letter, grid_offsets['start'])}{row}",
            "stop": f"{_offset(base_letter, grid_offsets['stop'])}{row}",
            "hours": f"{_offset(base_letter, grid_offsets['hours'])}{row}",
            "flag": f"{_offset(base_letter, grid_offsets['flag'])}{row}",
        })
    # Row 28's base column holds the static "Year" label; the actual
    # value cell is one column over (verified against the template: A28
    # = "Year" (label), B28 = the year value, same +1 pattern for every
    # block). Writing to the base column would clobber the label instead.
    year_cell = f"{_offset(base_letter, 1)}{YEAR_TOTAL_ROW}"
    # This column's "16. Total Hours" value cell shares the Hours column's
    # letter, one row down (verified: column C's is U28, column D's AC28 —
    # both match this same offset rule).
    total_hours_cell = f"{_offset(base_letter, grid_offsets['hours'])}{YEAR_TOTAL_ROW}"

    return {"meta": meta_cells, "grid": grid, "year": year_cell, "total_hours": total_hours_cell}


COLUMN_MAPS = {col: _build_column_map(base) for col, base in BLOCK_BASE_COL.items()}


def _set(ws, coord, value):
    cell = ws[coord]
    cell.value = value
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=FONT_BOLD)


def _set_font_only(ws, coord):
    """Style a cell without touching its value — for the Hours/Total Hours/
    item-17 cells, which are live formulas we must never overwrite, but
    which otherwise keep the template's own (differently-sized) font,
    clashing with the Start/Stop cells right next to them.
    """
    ws[coord].font = Font(name=FONT_NAME, size=FONT_SIZE, bold=FONT_BOLD)


def _plan_column_assignments(groups, rows_by_group):
    assignments = []
    for group_name, group in groups.items():
        remaining = rows_by_group[group_name]
        while remaining:
            chunk, remaining = remaining[:GRID_ROWS], remaining[GRID_ROWS:]
            assignments.append({"group": group_name, "meta": group["meta"], "rows": chunk})
    return assignments


def fill(paystub, groups, reserved, template_path, profile, out_path):
    """Mirrors of288_filler.fill()'s shape (paystub, groups, reserved, ...,
    out_path) but writes an .xlsx instead of a PDF. Caps at 4 columns per
    sheet copy (matching the template's own layout); a pay period needing
    more than that spills onto "..._page2.xlsx" etc., same convention as
    the PDF version.
    """
    rows_by_group = allocate_rows(paystub, groups, reserved)
    assignments = _plan_column_assignments(groups, rows_by_group)
    pages = [assignments[i:i + 4] for i in range(0, len(assignments), 4)] or [[]]

    out_paths = []
    grand_total = 0.0
    base, ext = (out_path.rsplit(".", 1) if "." in out_path else (out_path, "xlsx"))

    for page_num, page_assignments in enumerate(pages, start=1):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        _set(ws, TOP_CELLS["1_hired_at"], profile.get("1_hired_at", ""))
        _set(ws, TOP_CELLS["2_employee_common_identifier"], profile.get("2_employee_common_identifier", ""))
        _set(ws, TOP_CELLS["4_hiring_unit_name"], profile.get("4_hiring_unit_name", ""))
        _set(ws, TOP_CELLS["5_name"], paystub.employee_name.title())
        _set(ws, TOP_CELLS["6_hiring_unit_phone"], profile.get("6_hiring_unit_phone", ""))
        _set(ws, TOP_CELLS["7_hiring_unit_fax"], profile.get("7_hiring_unit_fax", ""))

        employment_type = profile.get("3_type_of_employment") or "Federal"
        type_cell = TYPE_OF_EMPLOYMENT_CELLS.get(employment_type, TYPE_OF_EMPLOYMENT_CELLS["Federal"])
        # The template ships with the Federal checkbox (M4) pre-marked "X"
        # (a leftover from whatever filled form it was derived from) —
        # clear every option first so Casual/Other don't end up with two
        # X's on the printed form.
        for cell in TYPE_OF_EMPLOYMENT_CELLS.values():
            _set(ws, cell, None)
        _set(ws, type_cell, "X")

        # Row 28's Year cell is a per-page constant, not per-incident data,
        # and the template ships with a hardcoded "2026" under every
        # column block (another leftover) — write the real year to all
        # four up front so an unused column never shows a stale year.
        for cmap in COLUMN_MAPS.values():
            _set(ws, cmap["year"], paystub.year)
            # Hours/Total Hours are live formulas (never written above),
            # but still need the same font as everything else so the grid
            # doesn't visually clash — applied to all four column blocks,
            # used or not, so an unused column doesn't stand out either.
            for field_row in cmap["grid"]:
                _set_font_only(ws, field_row["hours"])
            _set_font_only(ws, cmap["total_hours"])
        _set_font_only(ws, ITEM17_TOTAL_CELL)

        page_total = 0.0
        first_col_on_page = {}

        for col, assignment in zip(COLUMN_LETTERS, page_assignments):
            group_name = assignment["group"]
            cmap = COLUMN_MAPS[col]

            if group_name not in first_col_on_page:
                first_col_on_page[group_name] = col
                meta = assignment["meta"]
                _set(ws, cmap["meta"]["incident_name"], meta.get("incident_name", ""))
                _set(ws, cmap["meta"]["incident_order_number"], meta.get("incident_order_number", ""))
                _set(ws, cmap["meta"]["fire_code"], meta.get("fire_code", ""))
                _set(ws, cmap["meta"]["resource_request_number"], meta.get("resource_request_number", ""))
                _set(ws, cmap["meta"]["position_code"], meta.get("position_code", ""))
                _set(ws, cmap["meta"]["accounting_code"], meta.get("accounting_code", ""))
                if employment_type == "Casual":
                    _set(ws, cmap["meta"]["ad_class"], meta.get("ad_class", ""))
                    _set(ws, cmap["meta"]["ad_rate"], meta.get("ad_rate", ""))
            else:
                ref_col = first_col_on_page[group_name]
                _set(ws, SAME_AS_CHECKBOXES[col][ref_col], "X")

            col_total = 0.0
            for field_row, (d, start, stop, hours, flag_suffix) in zip(cmap["grid"], assignment["rows"]):
                _set(ws, field_row["mo"], d.month)
                _set(ws, field_row["day"], d.day)
                _set(ws, field_row["start"], start)
                _set(ws, field_row["stop"], stop)
                if flag_suffix:
                    _set(ws, field_row["flag"], flag_suffix)
                col_total += hours
            # Hours cells and this column's Total Hours cell are left
            # untouched — they're live formulas in the template (Hours is
            # computed from Start/Stop, Total sums the Hours column) and
            # will show the correct number once opened in a real
            # spreadsheet app. col_total here is only used for our own
            # returned grand_total, not written anywhere.

            page_total += col_total

        # Item 17 (grand total) is also a live formula in the template
        # (sums all four columns' Total Hours cells) — left untouched.

        page_path = out_path if page_num == 1 else f"{base}_page{page_num}.{ext}"
        wb.save(page_path)
        out_paths.append(page_path)
        grand_total += page_total

    return out_paths, grand_total
